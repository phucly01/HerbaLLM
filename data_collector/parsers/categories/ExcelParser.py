

from io import BytesIO
import openpyxl
from data_collector.parsers.categories.Category import Category
from data_collector.storage.Storage import Storage


class ExcelParser(Category):
    
    def process(self, bio:BytesIO, data_selector:list, storage:Storage)  -> bool:
        workbook = openpyxl.load_workbook(bio)
        sheets =  workbook.sheetnames
        for sheetname in sheets:
            selector_found = None
            for selector in data_selector:
                if sheetname in selector.values():
                    selector_found = selector
            if not selector_found:
                print(f'Nothing found for sheet {sheetname}.  Skipping ...')
                continue
            
            print(f'### Process sheet {sheetname}')
            sheet = workbook[sheetname]
            headers = [cell.value for cell in sheet[1]]
            print(f'### Saving to {storage.get_db_name()}:{storage.get_table_name()}')
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = dict(zip(headers, values))
                row['_id'] = row.pop(selector_found['id-map'])
                if storage.save([row]):
                    print(row)
                else:
                    return False
        return True